import shutil, os, time, logging, openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side
from datetime import datetime
from dotenv import load_dotenv

from trataExtrato import processaBradesco, processaBradesco_csv, processaItau, processaBB, processaBB_csv, processaOutros

def get_maximum_rows(sheet, col=None):
    if col is None:
        rows = 0
        for rows, row in enumerate(sheet, 1):
            if not all(col.value is None for col in row):
                rows += 1
    else:
        for i in range(1, 20000):
            if sheet.cell(row=i, column = col).value == None:
                rows = i 
                break
    return rows

def fProcNovos():
    #Criando um logger para o modulo
    module_logger = logging.getLogger('root.processaNovos')

    load_dotenv('./.env')

    oldAdress = r'.' #inicializando pasta origem no diretório do projeto
    newAdress = r'.' #inicializando pasta destino no diretório do projeto
    oldAdress += os.environ['OLD_ADDRESS']
    newAdress += os.environ['NEW_ADDRESS']
    destino2 = os.environ['POS_PROC2'] # Arquivo de destino das infos. Deve estar na raiz do projeto

    module_logger.info('Processando NOVOS ARQUIVOS de %s para %s', oldAdress, newAdress)
    module_logger.debug('Processando NOVOS ARQUIVOS de %s para %s', oldAdress, newAdress)

    lista = os.listdir(oldAdress) #lista separando apenas os arquivos do caminho.

    # *** lista_len recebe o tamanho da lista ***
    lista_len = len(lista)
    x = 0

    fill_pattern_Itau = PatternFill(patternType='solid', fgColor='B98E8E')
    fill_pattern_Bradesco = PatternFill(patternType='solid', fgColor='D2C4C4')
    fill_pattern_BB = PatternFill(patternType='solid', fgColor='9E5E5E')
    thin = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="ff0000")
    
    try:
        wb = openpyxl.load_workbook(destino2)
        folhaOriginais = wb['Linhas Originais']
        folhaPadronizadas = wb['Linhas Padronizadas']
        # max_FO = get_maximum_rows(folhaOriginais)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        folhaOriginais = wb.create_sheet(title='Linhas Originais')
        folhaOriginais['A1'] = r'Itaú'
        folhaOriginais.merge_cells('A1:E1')
        folhaOriginais['A1'].alignment = Alignment(horizontal='center', vertical='center')
        folhaOriginais['A2'] = r'data'
        folhaOriginais['B2'] = 'lançamento'
        folhaOriginais['C2'] = r'ag./origem'
        folhaOriginais['D2'] = r'valor (R$)'
        folhaOriginais['E2'] = r'OBS'
        for row in folhaOriginais['A1:E2']:
            for cell in row:
                cell.fill = fill_pattern_Itau
                cell.border = Border(top=double, left=thin, right=thin, bottom=double)

        folhaOriginais['G1'] = r'BRADESCO - Extrato de: Ag: 3261 | Conta: 43570-8'
        folhaOriginais.merge_cells('G1:N1')
        folhaOriginais['G1'].alignment = Alignment(horizontal='center', vertical='center')
        folhaOriginais['G2'] = r'Data'
        folhaOriginais['H2'] = r'Histórico'
        folhaOriginais['I2'] = r'Docto.'
        folhaOriginais['J2'] = r'Crédito'
        folhaOriginais['K2'] = r'Débito'
        folhaOriginais['L2'] = r'Saldo (R$)'
        folhaOriginais['M2'] = r'C/D'
        folhaOriginais['N2'] = r'OBS'
        for row in folhaOriginais['G1:N2']:
            for cell in row:
                cell.fill = fill_pattern_Bradesco
                cell.border = Border(top=double, left=thin, right=thin, bottom=double)

        folhaOriginais['P1'] = r'BB - AG4344-3 CC605705-5 - Atendimento: 4004-0001'
        folhaOriginais.merge_cells('P1:U1')
        folhaOriginais['P1'].alignment = Alignment(horizontal='center', vertical='center')
        folhaOriginais['P2'] = r'Data'
        folhaOriginais['Q2'] = r'Dependência Origem'
        folhaOriginais['R2'] = r'Histórico'
        folhaOriginais['S2'] = r'Data do Balancete'
        folhaOriginais['T2'] = r'Número do Documento'
        folhaOriginais['U2'] = r'Valor'
        for row in folhaOriginais['P1:S2']:
            for cell in row:
                cell.fill = fill_pattern_BB
                cell.border = Border(top=double, left=thin, right=thin, bottom=double)

        folhaPadronizadas = wb.create_sheet(title='Linhas Padronizadas')
        folhaPadronizadas['A1'] = r'Banco'
        folhaPadronizadas['B1'] = r'Mes'
        folhaPadronizadas['C1'] = r'Dia'
        folhaPadronizadas['D1'] = r'Data'
        folhaPadronizadas['E1'] = r'Descricao 1'
        folhaPadronizadas['F1'] = r'Descricao 2'
        folhaPadronizadas['G1'] = r'Valor'
        folhaPadronizadas['H1'] = r'Categorizacao1'
        # max_FO = get_maximum_rows(folhaOriginais)
        module_logger.debug('except')
    except Exception as inst:
        print(type(inst))    # the exception instance
        print(inst.args)     # arguments stored in .args
        print(inst) 

    max_FO = get_maximum_rows(folhaOriginais)



    while x < lista_len:
        module_logger.debug('Processando o nome %s', lista[x])
        dtProcessamento = datetime.now() # current date and time
        dtStrProcessamento = dtProcessamento.strftime("%Y%m%d-%H%M%S")
        caminhoCompleto_old = oldAdress + lista[x] #variável recebe caminho + arquivo, conforme indice
        caminhoCompleto_new = newAdress + dtStrProcessamento  + ' - ' + lista[x]  #variável recebe caminho + arquivo, conforme indice

        if ((lista[x].find('Bradesco')>=0) and (lista[x].find('.xls')>=0)):
            module_logger.debug('Entrei no if para Bradesco')
            novasLinhas = processaBradesco(caminhoCompleto_old)
            print ('As novas linhas sao:')
            print(novasLinhas)
            max_FOB = max_FO
            colInicioBradesco=6
            
            for linha in range(1, len(novasLinhas)):
                for col in range(len(novasLinhas[linha])):
                    folhaOriginais.cell(row=max_FOB, column=colInicioBradesco+col+1).value = novasLinhas[linha][col]
                    folhaOriginais.cell(row=max_FOB, column=colInicioBradesco+col+1).fill = fill_pattern_Bradesco
                    folhaOriginais.cell(row=max_FOB, column=colInicioBradesco+col+1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                max_FOB += 1

        elif ((lista[x].find('Bradesco')>=0) and (lista[x].find('.csv')>=0)):
            module_logger.debug('Entrei no if do .csv para Bradesco')
            novasLinhas = processaBradesco_csv(caminhoCompleto_old)
            print ('As novas linhas sao:')
            print(novasLinhas)
            max_FOI = max_FO
            colInicioBB=15

            for linha in range(1, len(novasLinhas)):
                for col in range(len(novasLinhas[linha])):
                    folhaOriginais.cell(row=max_FOI, column=colInicioBB+col+1).value = novasLinhas[linha][col]
                    folhaOriginais.cell(row=max_FOI, column=colInicioBB+col+1).fill = fill_pattern_BB
                    folhaOriginais.cell(row=max_FOI, column=colInicioBB+col+1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                max_FOI += 1

        elif (lista[x].find('Itau')>=0):
            module_logger.debug('Entrei no if para Itaú')
            novasLinhas = processaItau(caminhoCompleto_old)
            print ('As novas linhas sao:')
            print(novasLinhas)
            max_FOI = max_FO
            colInicioItau=0

            for linha in range(1, len(novasLinhas)):
                for col in range(len(novasLinhas[linha])):
                    folhaOriginais.cell(row=max_FOI, column=colInicioItau+col+1).value = novasLinhas[linha][col]
                    folhaOriginais.cell(row=max_FOI, column=colInicioItau+col+1).fill = fill_pattern_Itau
                    folhaOriginais.cell(row=max_FOI, column=colInicioItau+col+1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                max_FOI += 1

        elif ((lista[x].find('BB')>=0) and (lista[x].find('.pdf')>=0)):
            module_logger.debug('Entrei no if do .pdf para Banco do Brasil')
            novasLinhas = processaBB(caminhoCompleto_old)
            print ('As novas linhas sao:')
            print(novasLinhas)
            max_FOI = max_FO
            colInicioBB=15

            for linha in range(1, len(novasLinhas)-1):
                for col in range(len(novasLinhas[linha])):
                    folhaOriginais.cell(row=max_FOI, column=colInicioBB+col+1).value = novasLinhas[linha][col]
                    folhaOriginais.cell(row=max_FOI, column=colInicioBB+col+1).fill = fill_pattern_BB
                    folhaOriginais.cell(row=max_FOI, column=colInicioBB+col+1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                max_FOI += 1

        elif ((lista[x].find('BB')>=0) and (lista[x].find('.csv')>=0)):
            module_logger.debug('Entrei no if do .csv para Banco do Brasil')
            novasLinhas = processaBB_csv(caminhoCompleto_old)
            print ('As novas linhas sao:')
            print(novasLinhas)
            max_FOI = max_FO
            colInicioBB=15

            for linha in range(1, len(novasLinhas)):
                for col in range(len(novasLinhas[linha])):
                    folhaOriginais.cell(row=max_FOI, column=colInicioBB+col+1).value = novasLinhas[linha][col]
                    folhaOriginais.cell(row=max_FOI, column=colInicioBB+col+1).fill = fill_pattern_BB
                    folhaOriginais.cell(row=max_FOI, column=colInicioBB+col+1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                max_FOI += 1

        else:
            module_logger.debug('Entrei no if para outros arquivos')
            processaOutros(caminhoCompleto_old)

        shutil.move(caminhoCompleto_old, caminhoCompleto_new) #módulo 'shutil.move()' move os arquivos
        module_logger.debug('Processamento de %s para %s', caminhoCompleto_old, caminhoCompleto_new)
        x += 1
        time.sleep(2.5)

    # Salvando arquivo excell com todas as linhas novas
    wb.save(destino2)

    module_logger.info('NOVOS ARQUIVOS. Fim do processamento de %i arquivo(s)', x)


if __name__ == '__main__':
    fProcNovos()
    # main('Rebuild')
